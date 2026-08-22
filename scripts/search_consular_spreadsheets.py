from pathlib import Path
import re
from openpyxl import load_workbook

base = Path('/home/ubuntu/pesquisa-ancestralidade-jonathan-git/sources')
patterns = re.compile(r'schell|schnell|schel|shell|nicolaus|nikolaus|nicolau|johann|carlos|karl|becker', re.I)
outputs = []

for path in sorted(base.glob('konsulatsmatrikel-*.xlsx')):
    outputs.append(f'\n===== {path.name} =====\n')
    workbook = load_workbook(path, read_only=True, data_only=True)
    found = 0
    for sheet in workbook.worksheets:
        outputs.append(f'-- planilha: {sheet.title} --\n')
        for row_no, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = ['' if value is None else str(value) for value in row]
            text = ' | '.join(values)
            if patterns.search(text):
                outputs.append(f'linha {row_no}: {text}\n')
                found += 1
    if found == 0:
        outputs.append('Nenhuma linha com variante pesquisada.\n')

report = base / 'matriculas_planilhas_busca_variantes.txt'
report.write_text('Busca de variantes em planilhas de matrículas consulares\n'
                  'Padrão: Schell|Schnell|Schel|Shell|Nicolaus|Nikolaus|Nicolau|Johann|Carlos|Karl|Becker\n'
                  + ''.join(outputs), encoding='utf-8')
print(report)
